import openpyxl


class Bulletin:
    def __init__(self):
        self.choice = 0
        self.students = []
        self.subjects = ['Francais', "Mathematics", "English", "TOO", "Base de Donnees"]
        self.statistics = {}
        self.run_program()

    def print_menu(self):
        print("\nSelectionner une operation par son numero")
        print("1 - Entrer les donnees des etudiants(es)")
        print("2 - Afficher les statistiques de la classe")
        print("3 - Rechercher un etudiant(e) et afficher son bulletin")
        print("4 - Quitter")
        self.choice = int(input(">: "))

    def run_program(self):
        while True:
            self.print_menu()
            switch = {
                1: self.get_students_data,
                2: self.print_class_statistics,
                3: self.find_student,
                4: exit
            }
            switch.get(self.choice, self.print_menu)()

    def get_students_data(self):
        print("Vous pouvez ici entrer tous les étudiants, ou entrez 0 si vous finissez d'entrer")
        cpt = 1
        while True:
            student = {}
            print(f"Pour l'étudiant {cpt}:")
            full_name = input("Entrez le prénom suivi du nom ou tapez 0 pour finir: ")

            if full_name == '0':
                break
            first_name, last_name = full_name.split()
            student['first_name'] = first_name
            student['last_name'] = last_name

            student['grades'] = {}
            for subject in self.subjects:
                while True:
                    notes = input(f"Entrer la ou les notes /20 de {subject} (separées par un espace) : ")
                    notes_list = notes.split()

                    if len(notes_list) == 1:
                        note1 = int(notes_list[0])
                        if 0 <= note1 <= 20:
                            student['grades'][subject] = [note1, 0, note1]
                            break
                    elif len(notes_list) == 2:
                        note1, note2 = map(int, notes_list)
                        average = (note1 + note2) / 2
                        student['grades'][subject] = [note1, note2, average]

                        if 0 <= note1 <= 20 and 0 <= note2 <= 20:
                            break
            self.students.append(student)
            cpt += 1

        self.students.sort(key=lambda x: x['last_name'])
        self.calculate_averages()

    @staticmethod
    def get_mention(average):
        if average >= 16:
            mention = "Très bien"
        elif average >= 14:
            mention = "Bien"
        elif average >= 12:
            mention = "Assez bien"
        elif average >= 10:
            mention = "Passable"
        else:
            mention = "Insuffisant"
        return mention

    def calculate_averages(self):
        for i, student in enumerate(self.students):
            total = 0

            for subject in student.get('grades').values():
                total += subject[2]

            average = total / len(self.subjects)
            self.students[i]['average'] = average
            self.students[i]['mention'] = self.get_mention(average)
        self.calculate_statistics()
        self.generate_report()

    def calculate_statistics(self):
        total, highest, lowest = 0, 0, self.students[0]['average']

        for i, student in enumerate(self.students):
            average = student['average']
            total += average
            if average > highest:
                highest = average
            if average < lowest:
                lowest = average

            rank = 1
            for other_student in self.students:
                if other_student['average'] > average:
                    rank += 1
            self.students[i]['rank'] = rank

        self.statistics['highest'] = highest
        self.statistics['lowest'] = lowest
        self.statistics['class'] = total / len(self.students)

    def generate_report(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        sheet['A1'] = "Nom"
        sheet['B1'] = "Prénom"
        for i, subject in enumerate(self.subjects):
            sheet.cell(row=1, column=i + 3, value=subject)
        sheet.cell(row=1, column=len(self.subjects) + 3, value="Moyenne")
        sheet.cell(row=1, column=len(self.subjects) + 4, value="Mention")
        sheet.cell(row=1, column=len(self.subjects) + 5, value="Rang")

        for i, student in enumerate(self.students):
            row = i + 2
            sheet.cell(row=row, column=1, value=student['last_name'])
            sheet.cell(row=row, column=2, value=student['first_name'])
            for j, grade in enumerate(student['grades'].values()):
                sheet.cell(row=row, column=j + 3, value=grade[2])
            sheet.cell(row=row, column=j + 4, value=student['average'])
            sheet.cell(row=row, column=j + 5, value=student['mention'])
            sheet.cell(row=row, column=j + 6, value=student['rank'])

        row += 2
        sheet.cell(row=row, column=1, value="Moyenne la plus élevée:")
        sheet.cell(row=row, column=2, value=self.statistics['highest'])
        row += 1
        sheet.cell(row=row, column=1, value="Moyenne la plus faible:")
        sheet.cell(row=row, column=2, value=self.statistics['lowest'])
        row += 1
        sheet.cell(row=row, column=1, value="Moyenne de la classe:")
        sheet.cell(row=row, column=2, value=self.statistics['class'])

        workbook.save("rapport_de_la_classe.xlsx")

    def find_student(self):
        name = input("Entrez le prenom suivi du nom de l'étudiant à rechercher: ")
        found = False
        for student in self.students:
            originalName = f"{student['first_name']} {student['last_name']}"
            if name == originalName:
                print("\033[1;32m\n{0}\n".format("*" * 68))
                print("*\t\t\t*\033[1;36mBULLETIN DE NOTE\033[1;32m*\t\t          *\n")
                print("{0}\n\n".format("*" * 68))
                print("Nom et Prenom: {0}\t\t\t".format(student['last_name'] + " " + student['first_name']))
                print("Matricule: 12412923\n")
                print("Phone number: 95-18-10-19\t\t\t")
                print("Semestre: S2\n\n")
                print("Matieres\t\tNote 1\t\tNote 2\t\tMoyenne\n\n\n")

                total = 0
                for subject in self.subjects:
                    note1, note2, average = student['grades'][subject]
                    print("{0} \t\t {1:.1f}\t\t{2:.1f}\t\t{3:.2f}\n\n".format(subject, note1, note2, average))
                    total += average

                print("{0}\n".format("*" * 68))
                print("Total:\t{0:.2f}\n\n".format(total))
                print("Moyenne:\t{0:.2f}\n\n".format(student['average']))
                print("Mention:\t{0}\n\n".format(student['mention']))
                print("{0}\n".format("*" * 68))
                print("Rang:\t{0}\n".format(student['rank']))
                print("{0}\033[0m\n\n".format("*" * 68))
                found = True
                break

        if not found:
            print("\x1b[1;31mEtudiant introuvable.\x1b[0m\n")

    def print_class_statistics(self):
        print(f"Moyenne la plus élevée: {self.statistics.get('highest')}")
        print(f"Moyenne la plus faible: {self.statistics.get('lowest')}")
        print(f"Moyenne de la classe: {self.statistics.get('class')}")
